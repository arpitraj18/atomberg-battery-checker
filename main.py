import os
import random
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill

# Configuration
SIMULATION_MODE = os.getenv('SIMULATION_MODE', 'True').lower() == 'true'

report_data = []

def get_timestamp_days_ago(days):
    return int((datetime.now() - timedelta(days=days)).timestamp())

# --- MOCK DATA ---
mock_locks = [
    {'lock_id': 'L100', 'last_checked': get_timestamp_days_ago(0)},
    {'lock_id': 'L103', 'last_checked': get_timestamp_days_ago(29)},
    {'lock_id': 'L101', 'last_checked': get_timestamp_days_ago(60)}, # Stale
    {'lock_id': 'L102', 'last_checked': get_timestamp_days_ago(35)}, # Stale
    {'lock_id': 'L104', 'last_checked': get_timestamp_days_ago(31)}, # Stale
    {'lock_id': 'L105', 'last_checked': get_timestamp_days_ago(365)},# Stale
    {'lock_id': 'L106', 'last_checked': get_timestamp_days_ago(100)} # Orphan
]

mock_user_mapping = {
    'L100': {'user_id': 'U_A', 'fcm': 'token_A'},
    'L101': {'user_id': 'U_B', 'fcm': 'token_B'},
    'L102': {'user_id': 'U_C', 'fcm': 'token_C'},
    'L103': {'user_id': 'U_D', 'fcm': 'token_D'},
    'L104': {'user_id': 'U_E', 'fcm': 'token_E'},
    'L105': {'user_id': 'U_A', 'fcm': 'token_A'},
}

class DatabaseService:
    def get_stale_locks(self, days_threshold=30):
        cutoff = int((datetime.now() - timedelta(days=days_threshold)).timestamp())
        stale_locks = []
        
        if SIMULATION_MODE:
            print(f"[DynamoDB] Scanning for locks older than {days_threshold} days...")
            for lock in mock_locks:
                if lock['last_checked'] < cutoff:
                    stale_locks.append(lock['lock_id'])
        else:
            try:
                import boto3
                from boto3.dynamodb.conditions import Attr
                dynamodb = boto3.resource('dynamodb', region_name='us-east-1')
                table = dynamodb.Table('locks')
                response = table.scan(FilterExpression=Attr('last_checked').lt(cutoff))
                stale_locks = [item['lock_id'] for item in response.get('Items', [])]
            except Exception as e:
                print(f"DB Error: {e}")
                
        return stale_locks

    def get_user_details(self, lock_ids):
        users = []
        if SIMULATION_MODE:
            print(f"[Postgres] Fetching owners for locks: {lock_ids}")
            for lid in lock_ids:
                if lid in mock_user_mapping:
                    data = mock_user_mapping[lid]
                    users.append({
                        'lock_id': lid, 
                        'user_id': data['user_id'], 
                        'fcm_token': data['fcm']
                    })
                else:
                    self._log_orphan(lid)
        else:
            try:
                import psycopg2
                from psycopg2.extras import RealDictCursor
                conn = psycopg2.connect(
                    host=os.getenv('DB_HOST'), database=os.getenv('DB_NAME'),
                    user=os.getenv('DB_USER'), password=os.getenv('DB_PASSWORD')
                )
                cur = conn.cursor(cursor_factory=RealDictCursor)
                if lock_ids:
                    cur.execute("SELECT lock_id, user_id, fcm_id as fcm_token FROM lock_user_mapping WHERE lock_id IN %s", (tuple(lock_ids),))
                    users = cur.fetchall()
                cur.close()
                conn.close()
            except Exception as e:
                print(f"DB Error: {e}")
        return users

    def _log_orphan(self, lock_id):
        print(f"⚠️ Warning: Orphan Lock {lock_id}")
        report_data.append({
            'Timestamp': datetime.now().isoformat(),
            'Lock_ID': lock_id, 'User_ID': 'N/A', 'Status': 'FAILURE',
            'User_Action': 'N/A', 'Outcome': 'FAILED', 'Reason': 'Orphan Lock'
        })

def send_notifications(users, campaign_id):
    count = 0
    print(f"\nStarting Campaign: {campaign_id}")

    for user in users:
        success = False
        if SIMULATION_MODE:
            print(f"   ➡ [FCM] Sent to {user['user_id']} for lock {user['lock_id']}")
            success = True
        else:
            try:
                from firebase_admin import messaging, initialize_app
                try: initialize_app() 
                except ValueError: pass
                
                msg = messaging.Message(
                    token=user['fcm_token'],
                    data={'campaign_id': campaign_id},
                    notification=messaging.Notification(title='Battery Check', body='Please check battery')
                )
                messaging.send(msg)
                success = True
            except Exception as e:
                print(f"FCM Error: {e}")

        if success:
            count += 1
            report_data.append({
                'Timestamp': datetime.now().isoformat(),
                'Lock_ID': user['lock_id'], 'User_ID': user['user_id'],
                'Status': 'SUCCESS', 'User_Action': 'NO_RESPONSE',
                'Outcome': 'PENDING', 'Reason': f'Campaign: {campaign_id}'
            })
    return count

def simulate_analytics(campaign_id, users):
    clicks = []
    conversions = []
    
    if SIMULATION_MODE:
        print(f"\n🎲 [Analytics] Simulating interactions...")
        for user in users:
            # Random chance (50%) to click
            if random.random() < 0.5: 
                clicks.append({'User_ID': user['user_id']})
            
            # Random chance (30%) to convert
            if random.random() < 0.3: 
                conversions.append(user['user_id'])
                
        print(f"   ➡ Stats: {len(clicks)} clicks, {len(conversions)} checks.")
        
    return clicks, conversions

def generate_report(campaign_id, sent_count, clicks, conversions):
    filename = f"battery_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    clicked_ids = {c['User_ID'] for c in clicks}
    converted_ids = set(conversions)
    
    for row in report_data:
        if row['Status'] == 'SUCCESS':
            if row['User_ID'] in clicked_ids: 
                row['User_Action'] = 'OPENED'
            
            row['Outcome'] = 'EFFECTIVE' if row['User_ID'] in converted_ids else 'PENDING'

    df_logs = pd.DataFrame(report_data)
    cols = ['Timestamp', 'Lock_ID', 'User_ID', 'Status', 'User_Action', 'Outcome', 'Reason']
    df_logs = df_logs[[c for c in cols if c in df_logs.columns]]

    click_count = len(clicks)
    conv_count = len(conversions)
    ctr = (click_count / sent_count * 100) if sent_count else 0
    effectiveness = (conv_count / sent_count * 100) if sent_count else 0
    
    df_summary = pd.DataFrame([
        {'Metric': 'Campaign Date', 'Value': datetime.now().strftime('%Y-%m-%d')},
        {'Metric': 'Campaign ID', 'Value': campaign_id},
        {'Metric': 'Sent', 'Value': sent_count},
        {'Metric': 'Clicks (CTR)', 'Value': f"{click_count} ({ctr:.1f}%)"},
        {'Metric': 'Effectiveness', 'Value': f"{conv_count} ({effectiveness:.1f}%)"},
    ])

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_logs.to_excel(writer, index=False, sheet_name='Detailed Logs')
        df_summary.to_excel(writer, index=False, sheet_name='Dashboard')
        
        ws = writer.sheets['Detailed Logs']
        green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        yellow = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        for idx, row in enumerate(report_data, start=2):
            fill = None
            if row['Status'] == 'FAILURE': fill = red
            elif row['Status'] == 'SUCCESS':
                fill = green if row.get('User_Action') == 'OPENED' else yellow
            if fill:
                for col in range(1, len(df_logs.columns) + 1):
                    ws.cell(row=idx, column=col).fill = fill

    print(f"\n✅ Report generated: {filename}")

def main():
    mode = "SIMULATION" if SIMULATION_MODE else "PRODUCTION"
    print(f"--- Battery Check Job ({mode}) ---")
    
    db = DatabaseService()
    stale_locks = db.get_stale_locks()
    print(f" Found {len(stale_locks)} stale locks.")

    if stale_locks:
        users = db.get_user_details(stale_locks)
        campaign_id = f"bat_check_{datetime.now().strftime('%Y_W%U')}"
        
        # Single, clean function call
        sent_count = send_notifications(users, campaign_id)
        
        clicks, conversions = simulate_analytics(campaign_id, users)
        generate_report(campaign_id, sent_count, clicks, conversions)
    else:
        print("No stale locks found.")

if __name__ == "__main__":
    main()